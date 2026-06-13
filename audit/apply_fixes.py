import openpyxl, sys
sys.stdout.reconfigure(encoding="utf-8")

base = r"C:\Users\WALID\Projet ProClean AZ"

# ---------- FIX XL-005 : Tableau-de-Bord, facteur net réel 0,678 -> 0,6732 ----------
f1 = base + r"\Tableau-de-Bord-AZduDashboard.xlsx"  # placeholder, corrected below
f1 = base + r"\Tableau-de-Bord-AZduClean.xlsx"
wb = openpyxl.load_workbook(f1, data_only=False)
ws = wb["Pilotage_Mensuel"]
changed = 0
for row in ws.iter_rows():
    for cell in row:
        v = cell.value
        if isinstance(v, str):
            nv = v.replace("0.678", "0.6732").replace("67,8 %", "67,32 %").replace("67,8%", "67,32%")
            if nv != v:
                cell.value = nv
                changed += 1
                print(f"XL-005 {cell.coordinate}: {v!r} -> {nv!r}")
try:
    wb.calculation.fullCalcOnLoad = True
except Exception:
    pass
wb.save(f1)
print(f"XL-005: {changed} cellules modifiées, sauvegardé.\n")

# ---------- FIX XL-004 : Previsionnel, Scenarios divisions par MAX() -> IFERROR ----------
f2 = base + r"\Previsionnel_ProClean_AZ.xlsx"
wb2 = openpyxl.load_workbook(f2, data_only=False)
ws2 = wb2["Scenarios"]
targets = ["B13", "D13", "B14", "D14"]
g = 0
for coord in targets:
    c = ws2[coord]
    v = c.value
    if isinstance(v, str) and v.startswith("=") and "IFERROR" not in v.upper():
        inner = v[1:]  # strip leading =
        c.value = f"=IFERROR({inner},0)"
        g += 1
        print(f"XL-004 {coord}: {v} -> {c.value}")
try:
    wb2.calculation.fullCalcOnLoad = True
except Exception:
    pass
wb2.save(f2)
print(f"XL-004: {g} cellules protégées, sauvegardé.")
