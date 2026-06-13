# -*- coding: utf-8 -*-
"""Vérifie les 2 correctifs Excel après patch :
  XL-005  Tableau-de-Bord : facteur net réel 0,6732 (au lieu de 0,678)
  XL-004  Previsionnel Scenarios : plus de #DIV/0! quand plateau Airbnb/particuliers = 0
Et confirme que les chiffres certifiés du prévisionnel n'ont PAS bougé.
"""
import io, shutil, sys
from pathlib import Path
from openpyxl import load_workbook
import formulas

sys.stdout.reconfigure(encoding="utf-8")
BASE = Path(r"C:\Users\WALID\Projet ProClean AZ")
HERE = Path(__file__).parent
PREV = BASE / "Previsionnel_ProClean_AZ.xlsx"
TDB = BASE / "Tableau-de-Bord-AZduClean.xlsx"


def recalc(path):
    xl = formulas.ExcelModel().loads(str(path)).finish()
    return xl.calculate()


def val(sol, sheet, cell):
    needle = f"]{sheet.upper()}'!{cell.upper()}"
    for k, v in sol.items():
        if k.upper().endswith(needle):
            arr = getattr(v, "value", v)
            try:
                return arr[0][0]
            except Exception:
                return arr
    return None


def scan_errors(sol):
    bad = []
    for k, v in sol.items():
        arr = getattr(v, "value", v)
        flat = []
        try:
            for row in arr:
                for c in row:
                    flat.append(c)
        except Exception:
            flat = [arr]
        for c in flat:
            s = str(c)
            if s.startswith("#") and s.endswith(("!", "?")):
                bad.append((k, s))
    return bad


fails = []

# ---------- PREVISIONNEL au défaut : zéro erreur + chiffres certifiés ----------
print("=== Previsionnel (défaut) ===")
sol = recalc(PREV)
errs = scan_errors(sol)
print(f"  cellules en erreur : {len(errs)}")
if errs:
    for k, s in errs[:20]:
        print("   !", k, s)
    fails.append("Previsionnel défaut a des erreurs")

certified = {
    ("Resultat", "F5"): 36573,
    ("Resultat", "F18"): 22521,
    ("Resultat", "F20"): 2927,
    ("Scenarios", "C15"): 36573,   # CA réaliste = CA certifié (cohérence)
    ("Scenarios", "B19"): 13861,   # net pessimiste
    ("Scenarios", "C19"): 22521,   # net réaliste = net réel certifié
    ("Scenarios", "D19"): 31155,   # net optimiste
    ("Tresorerie", "B15"): 920,
}
for (sh, cl), exp in certified.items():
    got = val(sol, sh, cl)
    try:
        gotn = round(float(got))
    except Exception:
        gotn = got
    ok = gotn == exp
    print(f"  {sh}!{cl} = {gotn}  (attendu {exp})  {'OK' if ok else 'ÉCART'}")
    if not ok:
        fails.append(f"{sh}!{cl} {gotn}!={exp}")

# valeurs des 4 cellules protégées au défaut (doivent rester numériques, inchangées)
for cl in ("B13", "D13", "B14", "D14"):
    print(f"  Scenarios!{cl} = {val(sol,'Scenarios',cl)}")

# ---------- PREVISIONNEL plateau Airbnb=0 ET particuliers=0 : preuve anti #DIV/0! ----------
print("\n=== Previsionnel (Airbnb plateau=0 + particuliers plateau=0) ===")
WORK = HERE / "_verify_zero.xlsx"
shutil.copyfile(PREV, WORK)
wb = load_workbook(WORK)
pa = wb["Plan_Activite"]
# Mettre à zéro les plateaux Airbnb (ligne 13) et particuliers (ligne 16) B..M
zeroed = 0
for r in (13, 16):
    for col in "BCDEFGHIJKLM":
        c = pa[f"{col}{r}"]
        if isinstance(c.value, (int, float)):
            c.value = 0
            zeroed += 1
        elif isinstance(c.value, str) and c.value.startswith("="):
            # remplacer formule plateau par 0 pour forcer MAX=0
            c.value = 0
            zeroed += 1
wb.save(WORK)
print(f"  {zeroed} cellules plateau mises à 0")
sol2 = recalc(WORK)
errs2 = scan_errors(sol2)
div0 = [e for e in errs2 if "DIV" in e[1]]
print(f"  #DIV/0! restants : {len(div0)}   (toutes erreurs : {len(errs2)})")
for cl in ("B13", "D13", "B14", "D14"):
    print(f"  Scenarios!{cl} = {val(sol2,'Scenarios',cl)}")
if div0:
    for k, s in div0[:20]:
        print("   !", k, s)
    fails.append("DIV/0 persiste avec plateau=0")

# ---------- TABLEAU DE BORD : preuve du facteur 0,6732 ----------
print("\n=== Tableau-de-Bord : facteur net réel ===")
WORK2 = HERE / "_verify_tdb.xlsx"
shutil.copyfile(TDB, WORK2)
wb2 = load_workbook(WORK2)
pm = wb2["Pilotage_Mensuel"]
pm["B7"] = 1000  # CA réel mois 1 = 1000 €
wb2.save(WORK2)
sol3 = recalc(WORK2)
b22 = val(sol3, "Pilotage_Mensuel", "B22")
attendu = round(1000 * 0.6732 - 75)  # = 598
print(f"  B7=1000 -> B22 = {b22}  (attendu {attendu})  {'OK' if round(float(b22))==attendu else 'ÉCART'}")
if round(float(b22)) != attendu:
    fails.append(f"TDB B22 {b22}!={attendu}")
errs3 = scan_errors(sol3)
print(f"  erreurs TDB : {len(errs3)}")

# nettoyage
for w in (WORK, WORK2):
    try:
        w.unlink()
    except Exception:
        pass

print("\n" + ("=" * 40))
if fails:
    print("ÉCHEC :")
    for f in fails:
        print("  -", f)
    sys.exit(1)
print("TOUT VERT : 2 correctifs Excel validés, chiffres certifiés intacts, zéro #DIV/0!.")
