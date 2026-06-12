# -*- coding: utf-8 -*-
"""Preuve de parité Excel ↔ simulateur, de bout en bout.

1. Copie le prévisionnel certifié, change un jeu d'hypothèses NON-défaut
   (chaque écriture est vérifiée par le libellé de la ligne — le classeur vivant fait foi).
2. Recalcule TOUTES les formules avec le paquet `formulas` (pas d'Excel requis).
3. Extrait les résultats clés et les écrit dans parite_e2e_attendu.json.
4. Le test vitest src/lib/parite-e2e.test.ts rejoue les mêmes hypothèses dans
   computeModel() et exige les mêmes chiffres.
"""
import io
import json
import shutil
import sys
from pathlib import Path

from openpyxl import load_workbook

SRC = Path(r"C:\Users\WALID\Projet ProClean AZ\Previsionnel_ProClean_AZ.xlsx")
HERE = Path(__file__).parent
WORK = HERE / "parite_e2e.xlsx"
OUT = HERE / "parite_e2e_attendu.json"

# --- Jeu d'hypothèses NON-défaut (cellule, fragment de libellé attendu, valeur) ---
PATCHES = [
    ("B10", "Taux horaire B2B standard", 32),
    ("B11", "Part des contrats B2B en formule annuelle", 0.4),
    ("B21", "ACRE active la première année", "OUI"),
    ("B43", "Logiciel de facturation", 45),  # fixes : 15+12+15+45+8 = 95 €/mois
    ("B51", "Apport de départ", 1500),
    ("C8", "Airbnb — rotation", 32),  # 2,5 h × 32 €/h = 80 €/rotation
    ("B62", "Croissance du CA — année 2", 0.25),
]

shutil.copyfile(SRC, WORK)
wb = load_workbook(WORK)
ws = wb["Hypotheses"]
for cell, label_fragment, value in PATCHES:
    label = str(ws[f"A{cell[1:]}"].value or "")
    assert label_fragment.lower() in label.lower(), (
        f"Libellé inattendu en A{cell[1:]} : {label!r} (attendu : {label_fragment!r})"
    )
    ws[cell] = value
wb.save(WORK)
print("Patches appliqués (libellés vérifiés).", flush=True)

# --- Recalcul complet ---
import formulas  # noqa: E402

xl = formulas.ExcelModel().loads(str(WORK)).finish()
sol = xl.calculate()
print("Classeur recalculé.", flush=True)

# Index des solutions : clé "'[fichier]ONGLET'!CELLULE" -> valeur scalaire
def value(sheet: str, cell: str):
    needle = f"]{sheet.upper()}'!{cell.upper()}"
    for k, v in sol.items():
        if k.upper().endswith(needle):
            arr = getattr(v, "value", v)
            try:
                return arr[0][0]
            except Exception:
                return arr
    raise KeyError(f"{sheet}!{cell} introuvable dans la solution")


def num(sheet: str, cell: str) -> float:
    v = value(sheet, cell)
    return float(v)


expected = {
    "hypotheses_simulateur": {
        "hourlyB2B": 32,
        "annualShare": 0.4,
        "acre": True,
        "fixedMonthly": 95,
        "contribution": 1500,
        "airbnbPrice": 80,
        "growth": [0.25, 0.2, 0.1, 0.1],
    },
    "excel": {
        "ca_annee1": num("Resultat", "F5"),
        "net_reel": num("Resultat", "F18"),
        "net_mensuel_moyen": num("Resultat", "F19"),
        "croisiere": num("Resultat", "F20"),
        "point_bas": num("Tresorerie", "B15"),
        "apport_minimal": num("Tresorerie", "B17"),
        "apport_recommande": num("Tresorerie", "B18"),
        "scenario_pessimiste_ca": num("Scenarios", "B15"),
        "scenario_realiste_ca": num("Scenarios", "C15"),
        "scenario_optimiste_ca": num("Scenarios", "D15"),
        "scenario_pessimiste_net": num("Scenarios", "B19"),
        "scenario_realiste_net": num("Scenarios", "C19"),
        "scenario_optimiste_net": num("Scenarios", "D19"),
        "projection_ca": [num("Projection_5ans", f"{c}6") for c in "BCDEF"],
        "projection_net": [num("Projection_5ans", f"{c}11") for c in "BCDEF"],
    },
}

with io.open(OUT, "w", encoding="utf-8") as f:
    json.dump(expected, f, ensure_ascii=False, indent=2)
print(json.dumps(expected["excel"], ensure_ascii=False, indent=2))
print(f"\nOK -> {OUT}", flush=True)
sys.exit(0)
