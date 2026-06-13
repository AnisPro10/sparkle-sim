import openpyxl, sys
sys.stdout.reconfigure(encoding="utf-8")

base = r"C:\Users\WALID\Projet ProClean AZ"

# 1) Tableau de bord - Pilotage_Mensuel ligne 22
wb = openpyxl.load_workbook(base + r"\Tableau-de-Bord-AZduClean.xlsx", data_only=False)
print("=== Tableau-de-Bord sheets:", wb.sheetnames)
for sn in wb.sheetnames:
    ws = wb[sn]
    if "ilotage" in sn or "Pilotage" in sn:
        print(f"\n--- {sn} rows 18-24 ---")
        for r in range(18, 25):
            cells = []
            for c in range(1, 15):
                v = ws.cell(row=r, column=c).value
                if v is not None:
                    cells.append(f"{ws.cell(row=r, column=c).coordinate}={v!r}")
            if cells:
                print(f"r{r}: " + " | ".join(cells))

# search 0.678 anywhere in tableau de bord
print("\n=== search 0.678 / 0,678 in Tableau-de-Bord ===")
for sn in wb.sheetnames:
    ws = wb[sn]
    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if v is not None and ("0.678" in str(v) or "0,678" in str(v) or "67,8" in str(v) or "67.8" in str(v)):
                print(f"{sn}!{cell.coordinate} = {v!r}")

# 2) Previsionnel - Scenarios onglet
wb2 = openpyxl.load_workbook(base + r"\Previsionnel_ProClean_AZ.xlsx", data_only=False)
print("\n\n=== Previsionnel sheets:", wb2.sheetnames)
for sn in wb2.sheetnames:
    if "cenario" in sn or "Scenario" in sn or "Scénario" in sn:
        ws = wb2[sn]
        print(f"\n--- {sn} all formula cells with '/' ---")
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if isinstance(v, str) and v.startswith("=") and "/" in v:
                    print(f"{cell.coordinate} = {v}")
