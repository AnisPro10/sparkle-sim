# -*- coding: utf-8 -*-
"""Dump formules + valeurs des Excel certifiés -> vérité terrain pour la parité du simulateur."""
import io, json
from openpyxl import load_workbook

OUT = r"C:\Users\WALID\Projet ProClean AZ\sparkle-sim\audit"

def dump(path, sheets, dest):
    wbf = load_workbook(path)               # formules
    wbv = load_workbook(path, data_only=True)  # valeurs
    res = {}
    for name in sheets:
        if name not in wbf.sheetnames:
            res[name] = "ABSENT"
            continue
        sf, sv = wbf[name], wbv[name]
        rows = []
        for row in sf.iter_rows(min_row=1, max_row=min(sf.max_row, 120), max_col=min(sf.max_column, 20)):
            for c in row:
                if c.value is None:
                    continue
                v = sv.cell(row=c.row, column=c.column).value
                if isinstance(v, float):
                    v = round(v, 4)
                elif v is not None and not isinstance(v, (int, str, bool)):
                    v = str(v)
                rows.append({"cell": c.coordinate, "f": str(c.value)[:240], "v": v})
        res[name] = rows
    with io.open(dest, "w", encoding="utf-8") as f:
        json.dump(res, f, ensure_ascii=False, indent=1)
    print("OK", dest, {k: (len(v) if isinstance(v, list) else v) for k, v in res.items()})

dump(r"C:\Users\WALID\Projet ProClean AZ\Simulation_Juridique_ProClean.xlsx",
     ["Parametres", "Micro", "EI_reel", "EURL", "SASU", "Comparatif"],
     OUT + r"\verite_juridique.json")
dump(r"C:\Users\WALID\Projet ProClean AZ\Previsionnel_ProClean_AZ.xlsx",
     ["Hypotheses", "Resultat", "Tresorerie", "Scenarios", "Projection_5ans", "Ratios"],
     OUT + r"\verite_previsionnel.json")
